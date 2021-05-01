# Install the openpyxl library
from openpyxl import load_workbook
from yattag import Doc, indent
  
# Loading our Excel file
wb = load_workbook("./set_generator_xml_data.xlsx")
  
# creating the sheet 1 object
ws = wb.worksheets[0]

# Returning returns a triplet
doc, tag, text, line = Doc().ttl()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'
# xml_exchange 'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="http://download.autodesk.com/us/navisworks/schemas/nw-exchange-12.0.xsd" units="ft" filename="003.EXE.MOD.T01.R01_Andar Tipo_ Sem Sacada.rvt" filepath="G:\Drives compartilhados\Methodo Engenharia - Projetos\2. Propostas Fechadas\Procave - 10.2020\Fischer Dreams\02. Escopo\02.4 Escopo - BIM\02.4.1_FORMATO RVT\01_ORIGINAL"'

# Appends the String to document
doc.asis(xml_header)
# doc.asis(xml_schema)
  
# Iterating rows for getting the values of each row
# for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=14):
#     print([cell.value for cell in row])

with tag('exchange'):
    with tag('selectionsets'):
        with tag('viewfolder'):
            for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=14):
                row = [cell.value for cell in row]
                with tag('selectionset', name=str(row[1] + ' - ' + row[3])):
                    with tag('findspec', mode="all", disjoint="0"):
                        with tag('conditions'):
                            line('condition', '', test="equals", flags="10")
                            with tag('category'):
                                with tag('name', internal=row[4]):
                                    text(row[5])
                            with tag('property'):
                                with tag('name', internal=row[6]):
                                    text(row[7])
                            with tag('value'):
                                with tag('name', internal=row[8]):
                                    text(row[9])
                        line('locator', '')              
    
result = indent(
    doc.getvalue(),
    indentation='   ',
    indent_text=True
)
  
with open("set_generator_output.xml", "w") as f:
    f.write(result)